#!/usr/bin/env python
"""Auto-generate a vehicle logbook."""

import argparse
import datetime
import os
import random
import sys

import numpy

import openpyxl

TEMPLATE = os.path.join(os.path.dirname(__file__), "template.xlsx")
LOCATIONS = ("Cape Town CBD", "Rondebosch", "Claremont", "Sea Point",
             "Woodstock", "Belleville", "Milnerton", "Pinelands")
REASONS = {"Site Maintenance": 4, "Customer Meetings": 3, "Vendor Meetings": 1}


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--name", "-n", required=True)
    parser.add_argument("--tax-ref", "-t", required=True)
    parser.add_argument("--year", "-y", type=int,
                        default=datetime.date.today().year)
    parser.add_argument("--odometer-start", "-S", type=int, required=True)
    parser.add_argument("--odometer-end", "-E", type=int, required=True)
    parser.add_argument("--mileage", "-m", type=int, required=True)
    parser.add_argument("--make", required=True)
    parser.add_argument("--model", required=True)
    parser.add_argument("--model-year", required=True)
    parser.add_argument("--reg-number", required=True)
    parser.add_argument("--purchase-price", type=int, required=True)
    parser.add_argument("--output-file", "-o")
    args = parser.parse_args()
    assert args.odometer_end - args.odometer_start >= args.mileage
    return args


def daterange(year):
    """Get an iterator of dates in the tax year."""
    end_date = datetime.date(year, 3, 1)
    start_date = datetime.date(year - 1, 3, 1)
    period = end_date - start_date
    for n in range(period.days):
        yield start_date + datetime.timedelta(n)


def get_workbook(args):
    """Load the excel workbook from the template and set initial data."""
    wb = openpyxl.load_workbook(TEMPLATE)
    cover = wb["Cover"]
    cover["E13"] = args.name
    cover["E14"] = args.tax_ref
    cover["E15"] = args.year
    summary = wb["Summary"]
    summary["D7"] = args.make
    summary["D8"] = args.model
    summary["D9"] = args.model_year
    summary["D10"] = args.reg_number
    summary["D11"] = args.purchase_price
    return wb


def sampler(year, mileage):
    """Return a random value function based on the gamma distribution."""
    weekdays = len([d for d in daterange(year)
                    if d.weekday() in range(5)])
    shape = mileage / weekdays

    def func(day):
        if day.weekday() in range(5):
            x = round(numpy.random.gamma(shape))
        else:
            x = 0
        if x > 0:
            location = random.choice(LOCATIONS)
            reason = random.choice([r for r, weight in REASONS.items()
                                    for each in range(weight)])
            return location, reason, x
        else:
            return None, None, 0

    return func


def main():
    """Generate logbook."""
    args = parse_args()
    wb = get_workbook(args)
    log = wb["Log"]
    log["H7"] = args.odometer_start
    log["H378"] = args.odometer_end
    sample = sampler(args.year, args.mileage)
    days = daterange(args.year)
    total = 0
    for i, day in enumerate(days):
        row = i + 10
        location, reason, x = sample(day)
        total += x
        log.cell(row=row, column=2).value = day
        log.cell(row=row, column=3).value = f"{day.strftime('%A')}"
        log.cell(row=row, column=4).value = None
        log.cell(row=row, column=5).value = location
        log.cell(row=row, column=6).value = reason
        log.cell(row=row, column=8).value = x
    print(f"generated mileage: {total}")
    print(f"reported mileage: {args.mileage}")
    print(f"difference: {args.mileage - total}")
    if args.output_file:
        wb.save(args.output_file)
    return


if __name__ == "__main__":
    sys.exit(main())
